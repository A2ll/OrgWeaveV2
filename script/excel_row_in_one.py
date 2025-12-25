import os
from openpyxl import load_workbook
from xlsxwriter import Workbook
from loguru import logger

def optimize_excel_process(excel_path, start_col, end_col, sign_col):
    """
    用于处理excel文件行的的合并
    Args:
        excel_path (str): 需要处理的文件路径
        start_col (str): （大写字母）需要合并的起始列
        end_col (str): （大写字母）需要合并的结束列
        sign_col (str): （大写字母）用来判断的列名称

    Returns:
        str: 输出文件地址
    """
    # 列名转换（A->0）
    sign_idx = ord(sign_col.upper()) - 65
    start_idx = ord(start_col.upper()) - 65
    end_idx = ord(end_col.upper()) - 65

    # ==================== 阶段1：识别合并区域 ====================
    logger.info('识别合并区域')
    wb = load_workbook(excel_path, read_only=True)
    ws = wb.active
    
    # 读取所有标志列值（从第2行开始）
    #! sign_values = [row[sign_idx].value for row in ws.iter_rows(min_row=2, values_only=True)]
    sign_values = [row[sign_idx] for row in ws.iter_rows(min_row=2, values_only=True)]
    merged_regions = []
    
    if sign_values:
        current_val = sign_values[0]
        start_row = 2  # Excel实际行号
        
        for i in range(1, len(sign_values)+1):
            if i == len(sign_values) or sign_values[i] != current_val:
                end_row = i + 1  # Excel实际行号（i从0开始，对应行号2+i）
                #! if end_row - start_row > 1:
                if end_row - start_row >= 1:
                    merged_regions.append((start_row, end_row))
                if i < len(sign_values):
                    current_val = sign_values[i]
                    start_row = end_row + 1
    
    wb.close()
    
    # ==================== 阶段2：流式处理写入 ====================
    logger.info('流式处理写入')
    output_path = os.path.splitext(excel_path)[0] + "_optimized.xlsx"
    wb_writer = Workbook(output_path)
    ws_writer = wb_writer.add_worksheet()
    merge_format = wb_writer.add_format({'align': 'left', 'valign': 'vcenter'})
    
    # 读取原始数据流
    wb = load_workbook(excel_path, read_only=True)
    ws = wb.active
    
    # 处理标题行
    for col_idx, cell in enumerate(next(ws.iter_rows(min_row=1))):
        ws_writer.write(0, col_idx, cell.value)
    
    # 数据行处理
    current_region = 0
    for row_num, row in enumerate(ws.iter_rows(min_row=2), start=1):
        excel_row_num = row_num + 1  # Excel实际行号
        
        # 判断是否在合并区域
        if current_region < len(merged_regions):
            start_merge, end_merge = merged_regions[current_region]
            
            if excel_row_num == start_merge:
                # 处理合并起始行
                for col in range(start_idx, end_idx+1):
                    cell_value = row[col].value
                    ws_writer.merge_range(
                        first_row=row_num,
                        first_col=col,
                        last_row=row_num + (end_merge - start_merge),
                        last_col=col,
                        data=cell_value,
                        cell_format=merge_format
                    )
                
                # 写入其他列
                for col_idx, cell in enumerate(row):
                    if not (start_idx <= col_idx <= end_idx):
                        ws_writer.write(row_num, col_idx, cell.value)
                
                current_region += 1
            elif start_merge < excel_row_num <= end_merge:
                # 跳过合并区域的中间行
                for col_idx, cell in enumerate(row):
                    if not (start_idx <= col_idx <= end_idx):
                        ws_writer.write(row_num, col_idx, cell.value)
            else:
                # 正常写入
                for col_idx, cell in enumerate(row):
                    ws_writer.write(row_num, col_idx, cell.value)
        else:
            # 正常写入
            for col_idx, cell in enumerate(row):
                ws_writer.write(row_num, col_idx, cell.value)
    
    wb.close()
    wb_writer.close()
    logger.info('处理完成')
    return output_path

if __name__ == '__main__':
    file_path = r'xxxx.xlsx'
    optimize_excel_process(file_path, "A", "N", "B")

